import os
import logging
from datetime import datetime
from typing import Dict, List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReportGenerator:
    """Handles Excel report generation with formatting"""
    
    def __init__(self, output_directory: Optional[str] = None):
        """Initialize Excel report generator
        
        Args:
            output_directory: Directory to save reports (defaults to current directory)
        """
        self.output_directory = output_directory or os.getcwd()
    
    def create_campaign_report(self, df: pd.DataFrame, use_openai: bool = True, processing_stats: Optional[Dict] = None) -> str:
        """Create comprehensive Excel report with campaign descriptions and summary
        
        Args:
            df: DataFrame with campaign data
            use_openai: Whether OpenAI was used for generation
            processing_stats: Dictionary with processing statistics
            
        Returns:
            Path to the created report
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"campaign_report_{timestamp}.xlsx"
        output_path = os.path.join(self.output_directory, filename)
        
        # Process dataframe for display
        df_display = df.copy()
        
        # Enhanced column ordering - Raw SF data first, then AI content
        priority_sf_columns = [
            'Id', 'Name', 'Channel__c', 'Type', 'Status', 'IsActive', 'Description'
        ]
        
        additional_sf_columns = [
            'Sub_Channel__c', 'Sub_Channel_Detail__c', 'Integrated_Marketing__c',
            'Intended_Product__c', 'TCP_Campaign__c', 'TCP_Program__c', 'TCP_Theme__c',
            'Vendor__c', 'Vertical__c', 'Marketing_Message__c', 'Territory__c',
            'Intended_Country__c', 'Non_Attributable__c', 'Program__c',
            'Short_Description_for_Sales__c', 'Recent_Member_Count'
        ]
        
        ai_columns = ['AI_Prompt', 'AI_Sales_Description']
        
        # Create final column order
        final_column_order = []
        for col in priority_sf_columns:
            if col in df_display.columns:
                final_column_order.append(col)
        
        for col in additional_sf_columns:
            if col in df_display.columns and col not in final_column_order:
                final_column_order.append(col)
        
        # Add AI columns at the end (prompt before description)
        for col in ai_columns:
            if col in df_display.columns:
                final_column_order.append(col)
        
        # Reorder columns
        df_display = df_display[final_column_order]
        
        # Calculate comprehensive metrics for summary
        total_campaigns = len(df)
        campaigns_with_ai = df['AI_Sales_Description'].notna().sum() if 'AI_Sales_Description' in df.columns else 0
        campaigns_with_errors = df['AI_Sales_Description'].str.contains('Error generating', na=False).sum() if 'AI_Sales_Description' in df.columns else 0
        success_rate = (campaigns_with_ai / total_campaigns * 100) if total_campaigns > 0 else 0
        
        # Average description length (excluding errors)
        avg_desc_length = 0
        if 'AI_Sales_Description' in df.columns:
            descriptions = df['AI_Sales_Description'].dropna().tolist()
            valid_descriptions = [desc for desc in descriptions if isinstance(desc, str) and 'Error generating' not in desc]
            if valid_descriptions:
                avg_desc_length = sum(len(desc) for desc in valid_descriptions) / len(valid_descriptions)
        
        # Create enhanced summary DataFrame
        summary_data = {
            'Metric': [
                'Total Campaigns Queried',
                'Total Campaigns Processed',
                'Campaigns with AI Descriptions',
                'Campaigns with Processing Errors',
                'Processing Success Rate',
                'Average Description Length',
                'Total Campaign Members',
                'Processing Time (minutes)',
                'Unique Channels',
                'Unique Verticals',
                'Unique Sales Territories',
                'Campaigns with Attribution Tracking',
                'Sales Generated Campaigns',
                'Regular Marketing Campaigns',
                'Campaigns with Product Focus',
                'Processing Date'
            ],
            'Value': [
                processing_stats.get('total_campaigns_queried', 'N/A') if processing_stats else 'N/A',
                total_campaigns,
                campaigns_with_ai,
                campaigns_with_errors,
                f"{success_rate:.1f}%",
                f"{avg_desc_length:.1f} chars",
                processing_stats.get('total_members', 'N/A') if processing_stats else 'N/A',
                processing_stats.get('processing_time_minutes', 'N/A') if processing_stats else 'N/A',
                df['Channel__c'].nunique() if 'Channel__c' in df.columns else 0,
                df['Vertical__c'].nunique() if 'Vertical__c' in df.columns else 0,
                df['Territory__c'].nunique() if 'Territory__c' in df.columns else 0,
                df['Non_Attributable__c'].eq(False).sum() if 'Non_Attributable__c' in df.columns else 'N/A',
                df['Channel__c'].eq('Sales Generated').sum() if 'Channel__c' in df.columns else 0,
                df['Channel__c'].ne('Sales Generated').sum() if 'Channel__c' in df.columns else 0,
                df['Intended_Product__c'].ne('General').sum() if 'Intended_Product__c' in df.columns else 0,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        
        # Create Excel writer with multiple sheets
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Main campaign data sheet
            df_display.to_excel(writer, sheet_name='Campaign Data', index=False)
            
            # Processing summary sheet
            summary_df.to_excel(writer, sheet_name='Processing Summary', index=False)
            
            # Format all sheets with RingCentral branding
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # RingCentral color scheme
                rc_blue = "0684BC"  # RingCentral blue
                rc_dark_blue = "045A8D"  # Darker RingCentral blue
                white = "FFFFFF"
                
                # Format headers
                header_font = Font(bold=True, color=white)
                header_fill = PatternFill(start_color=rc_blue, end_color=rc_blue, fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(vertical='center')
                
                # Set standard row height for all rows
                for row in worksheet.iter_rows():
                    worksheet.row_dimensions[row[0].row].height = 15
                
                # For the main Campaign Data sheet, highlight AI columns
                if sheet_name == 'Campaign Data':
                    ai_col_fill = PatternFill(start_color=rc_dark_blue, end_color=rc_dark_blue, fill_type="solid")
                    ai_col_font = Font(bold=True, color=white)
                    for col_idx, col_name in enumerate(final_column_order, 1):
                        if col_name in ai_columns:
                            cell = worksheet.cell(row=1, column=col_idx)
                            cell.fill = ai_col_fill
                            cell.font = ai_col_font
                    
                    # Freeze panes at first row and key columns
                    worksheet.freeze_panes = 'E2'  # Freeze first 4 columns (Id, Name, Channel, Type) and header row
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # Set different widths for different column types
                    if sheet_name == 'Campaign Data' and column[0].value in ai_columns:
                        adjusted_width = min(max_length + 2, 80)  # AI columns can be wider
                    else:
                        adjusted_width = min(max_length + 2, 50)  # Standard columns
                    
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        logging.info(f"Comprehensive report saved to {output_path}")
        return output_path 

 